from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import io

from app.database import get_db
from app import models, schemas
from app.routers.auth import get_scoped_faculty_id, get_current_user
from app.activity_helper import log_activity

router = APIRouter()


def _calc_letter(total: float) -> tuple[str, float]:
    if total >= 97: return ("A+", 4.0)
    if total >= 93: return ("A",  4.0)
    if total >= 90: return ("A-", 3.7)
    if total >= 87: return ("B+", 3.3)
    if total >= 83: return ("B",  3.0)
    if total >= 80: return ("B-", 2.7)
    if total >= 77: return ("C+", 2.3)
    if total >= 73: return ("C",  2.0)
    if total >= 70: return ("C-", 1.7)
    if total >= 67: return ("D+", 1.3)
    if total >= 60: return ("D",  1.0)
    return ("F", 0.0)


# ── List ──────────────────────────────────────────────────────────────────────
@router.get("/")
def list_grades(
    student_id : Optional[str] = Query(None),
    course_id  : Optional[str] = Query(None),
    semester   : Optional[str] = Query(None),
    skip       : int           = Query(0, ge=0),
    limit      : int           = Query(100, ge=1, le=1000),
    db: Session = Depends(get_db),
    scoped_faculty_id: Optional[str] = Depends(get_scoped_faculty_id),
):
    q = db.query(models.Grade)
    if scoped_faculty_id: q = q.filter(models.Grade.faculty_id == scoped_faculty_id)
    if student_id:        q = q.filter(models.Grade.student_id == student_id)
    if course_id:         q = q.filter(models.Grade.course_id  == course_id)
    if semester:          q = q.filter(models.Grade.semester    == semester)
    grades = q.offset(skip).limit(limit).all()
    return [
        {'id': g.id, 'student_id': g.student_id, 'course_id': g.course_id,
         'faculty_id': g.faculty_id, 'semester': g.semester,
         'midterm': g.midterm, 'final_exam': g.final_exam,
         'assignments': g.assignments, 'oral': g.oral, 'practical': g.practical,
         'total': g.total, 'grade_letter': g.grade_letter, 'grade_points': g.grade_points}
        for g in grades
    ]


# ── Excel Export (MUST be before /{grade_id}) ─────────────────────────────────
@router.get("/export-excel")
def export_grades_excel(
    course_id  : Optional[str] = Query(None),
    semester   : Optional[str] = Query(None),
    faculty_id : Optional[str] = Query(None),
    db: Session = Depends(get_db),
    scoped_faculty_id: Optional[str] = Depends(get_scoped_faculty_id),
    user: models.User = Depends(get_current_user),
):
    """Export grades as formatted شيت كنترول Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl غير مثبت على الخادم")

    fid = scoped_faculty_id or faculty_id
    q = db.query(models.Grade)
    if fid:       q = q.filter(models.Grade.faculty_id == fid)
    if course_id: q = q.filter(models.Grade.course_id  == course_id)
    if semester:  q = q.filter(models.Grade.semester    == semester)
    grades = q.order_by(models.Grade.student_id).all()

    student_ids = list({g.student_id for g in grades})
    students = db.query(models.Student).filter(models.Student.student_id.in_(student_ids)).all()
    name_map = {s.student_id: s.name for s in students}

    course_name = course_id or "—"
    if course_id:
        c = db.get(models.Course, course_id)
        if c: course_name = f"{c.name} ({c.id})"

    wb = Workbook()
    ws = wb.active
    ws.title = "شيت الكنترول"
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill("solid", fgColor="1e3a5f")
    alt_fill    = PatternFill("solid", fgColor="f0f4f8")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font  = Font(bold=True, size=14, color="1e3a5f")
    thin   = Side(style="thin", color="aaaaaa")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right  = Alignment(horizontal="right",  vertical="center")

    ws.merge_cells("A1:K1")
    ws["A1"] = f"شيت الكنترول — {course_name}   |   الفصل: {semester or '—'}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    headers = [
        ("م", 5), ("كود الطالب", 14), ("اسم الطالب", 28),
        ("أعمال سنة\n(10)", 12), ("شفوي\n(10)", 12), ("ميد ترم\n(30)", 12),
        ("نصف سنة\n(50)", 12), ("نهائي\n(50)", 12), ("المجموع\n(100)", 12),
        ("التقدير", 10), ("النقاط", 10),
    ]
    for col, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 38

    for idx, g in enumerate(grades, 1):
        row = idx + 2
        fill = alt_fill if idx % 2 == 0 else None
        midterm     = float(g.midterm     or 0)
        assignments = float(g.assignments or 0)
        oral        = float(g.oral        or 0)
        half_year   = midterm + assignments + oral
        final       = float(g.final_exam  or 0)
        total       = float(g.total       or (half_year + final))
        letter      = g.grade_letter or _calc_letter(total)[0]
        points      = g.grade_points or _calc_letter(total)[1]

        vals = [idx, g.student_id, name_map.get(g.student_id, "—"),
                assignments or "—", oral or "—", midterm or "—",
                half_year or "—", final or "—", total, letter, points]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = border
            cell.alignment = center if col != 3 else right
            if fill: cell.fill = fill
            if col == 10:
                if letter == "F":         cell.font = Font(bold=True, color="cc0000")
                elif letter.startswith("A"): cell.font = Font(bold=True, color="006400")

    ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"grades_{course_id or 'all'}_{semester or 'all'}.xlsx".replace(" ", "_")
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ── Template download (MUST be before /{grade_id}) ────────────────────────────
@router.get("/template-excel")
def download_grade_template(
    course_id: Optional[str] = Query(None),
    semester:  Optional[str] = Query(None),
    user: models.User = Depends(get_current_user),
):
    """Download blank grade entry template."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl غير مثبت")

    wb = Workbook()
    ws = wb.active
    ws.title = "شيت الدرجات"
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill("solid", fgColor="1e3a5f")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="aaaaaa")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    headers = [
        ("كود الطالب", 15), ("اسم الطالب", 25),
        ("أعمال سنة", 12), ("شفوي", 12), ("ميد ترم", 12),
        ("نهائي", 12), ("المجموع", 12), ("التقدير", 10), ("النقاط", 10),
    ]
    for col, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    example = ["2025001", "مثال: محمد علي", 8, 7, 25, 45, 85, "B+", 3.0]
    for col, v in enumerate(example, 1):
        cell = ws.cell(row=2, column=col, value=v)
        cell.border = border; cell.alignment = center
        cell.font = Font(italic=True, color="888888")

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"grade_template_{course_id or 'course'}_{semester or 'sem'}.xlsx".replace(" ", "_")
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ── Get single (parameterized — MUST be after static routes) ──────────────────
@router.get("/{grade_id}")
def get_grade(
    grade_id: int,
    db: Session = Depends(get_db),
    scoped_faculty_id: Optional[str] = Depends(get_scoped_faculty_id),
):
    q = db.query(models.Grade).filter(models.Grade.id == grade_id)
    if scoped_faculty_id:
        q = q.filter(models.Grade.faculty_id == scoped_faculty_id)
    g = q.first()
    if not g:
        raise HTTPException(status_code=404, detail="Grade record not found or access denied")
    return g


# ── Create ────────────────────────────────────────────────────────────────────
@router.post("/")
def create_grade(
    data: schemas.GradeCreate,
    db: Session = Depends(get_db),
    scoped_faculty_id: Optional[str] = Depends(get_scoped_faculty_id),
    user: models.User = Depends(get_current_user),
):
    if scoped_faculty_id:
        data.faculty_id = scoped_faculty_id
    existing = db.query(models.Grade).filter(
        models.Grade.student_id == data.student_id,
        models.Grade.course_id  == data.course_id,
        models.Grade.semester   == data.semester
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail="Grade record already exists for this student/course/semester")
    grade = models.Grade(**data.model_dump())
    db.add(grade); db.commit(); db.refresh(grade)
    log_activity(db=db, user_id=user.id, faculty_id=scoped_faculty_id,
                 entity_type="grade", entity_id=str(grade.id), action="create",
                 description=f"Created grade for {data.student_id}: {data.course_id} - {data.total}")
    return grade


# ── Update ────────────────────────────────────────────────────────────────────
@router.put("/{grade_id}")
def update_grade(
    grade_id: int,
    data: schemas.GradeUpdate,
    db: Session = Depends(get_db),
    scoped_faculty_id: Optional[str] = Depends(get_scoped_faculty_id),
    user: models.User = Depends(get_current_user),
):
    q = db.query(models.Grade).filter(models.Grade.id == grade_id)
    if scoped_faculty_id:
        q = q.filter(models.Grade.faculty_id == scoped_faculty_id)
    g = q.first()
    if not g:
        raise HTTPException(status_code=404, detail="Grade record not found or access denied")
    update_data = data.model_dump(exclude_none=True)
    if scoped_faculty_id and "faculty_id" in update_data:
        del update_data["faculty_id"]
    for k, v in update_data.items():
        setattr(g, k, v)
    db.commit(); db.refresh(g)
    log_activity(db=db, user_id=user.id, faculty_id=scoped_faculty_id,
                 entity_type="grade", entity_id=str(grade_id), action="update",
                 description=f"Updated grade: {list(update_data.keys())}")
    return g


# ── Excel Import ──────────────────────────────────────────────────────────────
@router.post("/import-excel")
def import_grades_excel(
    file       : UploadFile = File(...),
    course_id  : str        = Query(...),
    semester   : str        = Query(...),
    faculty_id : Optional[str] = Query(None),
    db: Session = Depends(get_db),
    scoped_faculty_id: Optional[str] = Depends(get_scoped_faculty_id),
    user: models.User = Depends(get_current_user),
):
    """Import grades from Excel. Columns: كود الطالب | أعمال سنة | شفوي | ميد ترم | نهائي | المجموع | التقدير"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl غير مثبت على الخادم")

    fid = scoped_faculty_id or faculty_id
    if not (file.filename or "").endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="يجب رفع ملف Excel بامتداد .xlsx")

    content = file.file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    COL_MAP = {
        "كود الطالب": "student_id", "student_id": "student_id", "رقم الطالب": "student_id", "كود": "student_id",
        "أعمال سنة": "assignments", "assignments": "assignments", "أعمال": "assignments",
        "شفوي": "oral",         "oral": "oral",
        "ميد ترم": "midterm",   "midterm": "midterm",
        "نهائي": "final_exam",  "final_exam": "final_exam", "final": "final_exam",
        "المجموع": "total",     "total": "total",
        "التقدير": "grade_letter", "grade_letter": "grade_letter",
        "النقاط": "grade_points",  "grade_points": "grade_points",
    }

    header_row = None
    col_index: dict[str, int] = {}
    for r in ws.iter_rows():
        for cell in r:
            val = str(cell.value or "").strip()
            if val in ("كود الطالب", "student_id", "رقم الطالب", "كود"):
                header_row = cell.row; break
        if header_row: break

    if not header_row:
        raise HTTPException(status_code=400, detail="لم يتم العثور على رأس الجدول — تأكد من وجود عمود 'كود الطالب'")

    for cell in ws[header_row]:
        val = str(cell.value or "").strip()
        if val in COL_MAP:
            col_index[COL_MAP[val]] = cell.column

    if "student_id" not in col_index:
        raise HTTPException(status_code=400, detail="عمود 'كود الطالب' غير موجود")

    created = updated = skipped = 0
    errors: list[str] = []

    def _v(row, field):
        col = col_index.get(field)
        if col is None: return None
        v = row[col - 1].value
        if v is None or str(v).strip() in ("", "—", "-"): return None
        try: return float(v)
        except (TypeError, ValueError): return str(v).strip()

    for row in ws.iter_rows(min_row=header_row + 1):
        sid = str(row[col_index["student_id"] - 1].value or "").strip()
        if not sid or sid == "None": continue

        assignments  = _v(row, "assignments")
        oral         = _v(row, "oral")
        midterm      = _v(row, "midterm")
        final_exam   = _v(row, "final_exam")
        total        = _v(row, "total")
        grade_letter = _v(row, "grade_letter")
        grade_points = _v(row, "grade_points")

        if total is None:
            parts = [x for x in [assignments, oral, midterm, final_exam] if isinstance(x, (int, float))]
            total = sum(parts) if parts else None

        if total is not None and not grade_letter:
            grade_letter, grade_points = _calc_letter(float(total))

        try:
            existing = db.query(models.Grade).filter(
                models.Grade.student_id == sid,
                models.Grade.course_id  == course_id,
                models.Grade.semester   == semester,
            ).first()

            if existing:
                if assignments  is not None: existing.assignments  = float(assignments)
                if oral         is not None: existing.oral         = float(oral)
                if midterm      is not None: existing.midterm      = float(midterm)
                if final_exam   is not None: existing.final_exam   = float(final_exam)
                if total        is not None: existing.total        = float(total)
                if grade_letter is not None: existing.grade_letter = str(grade_letter)
                if grade_points is not None: existing.grade_points = float(grade_points)
                updated += 1
            else:
                def _f(v): return float(v) if isinstance(v, (int, float)) else None
                db.add(models.Grade(
                    student_id=sid, course_id=course_id,
                    semester=semester, faculty_id=fid,
                    assignments=_f(assignments), oral=_f(oral),
                    midterm=_f(midterm), final_exam=_f(final_exam),
                    total=_f(total),
                    grade_letter=str(grade_letter) if grade_letter else None,
                    grade_points=_f(grade_points),
                ))
                created += 1
        except Exception as ex:
            errors.append(f"طالب {sid}: {ex}"); skipped += 1

    db.commit()
    log_activity(db=db, user_id=user.id, faculty_id=fid,
                 entity_type="grade", entity_id=course_id, action="bulk_import",
                 description=f"Excel import: {course_id}/{semester} created={created} updated={updated} skipped={skipped}")
    return {"success": True, "course_id": course_id, "semester": semester,
            "created": created, "updated": updated, "skipped": skipped, "errors": errors[:10]}


# ── Delete ────────────────────────────────────────────────────────────────────
@router.delete("/{grade_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_grade(
    grade_id: int,
    db: Session = Depends(get_db),
    scoped_faculty_id: Optional[str] = Depends(get_scoped_faculty_id),
    user: models.User = Depends(get_current_user),
):
    q = db.query(models.Grade).filter(models.Grade.id == grade_id)
    if scoped_faculty_id:
        q = q.filter(models.Grade.faculty_id == scoped_faculty_id)
    g = q.first()
    if not g:
        raise HTTPException(status_code=404, detail="Grade record not found or access denied")
    db.delete(g); db.commit()
    log_activity(db=db, user_id=user.id, faculty_id=scoped_faculty_id,
                 entity_type="grade", entity_id=str(grade_id), action="delete",
                 description="Deleted grade record")
